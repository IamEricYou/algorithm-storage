import json
import openpyxl

jsonfile = 'jsons/kkday_prod_list.json'

with open(jsonfile) as json_file:
    data = json.load(json_file)

wb = openpyxl.load_workbook(filename='product.xlsx')
ws = wb.worksheets[0]
prods = data.get('prods')

row_index = 1

for p in prods:
    try:
        ws.cell(row=row_index, column=1).value = p.get('prod_no')
        ws.cell(row=row_index, column=2).value = p.get('prod_url_no')
        ws.cell(row=row_index, column=3).value = str(p.get('prod_name'))
        ws.cell(row=row_index, column=4).value = str(p.get('introduction'))
        row_index += 1
    except:
        print(p.get('prod_name'))

wb.save(filename='product.xlsx')
wb.close()