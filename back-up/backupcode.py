"""
엑셀 bootpay 주문완료건 작업
"""

import pandas as pd
import numpy as np
import openpyxl

from flanb_payments.models import ProductPayment
def check_value_is_not_nan(val):
    if type(val) == float and np.isnan(val):
        return False
    return True

def update_tracking_info():
    f = pd.read_excel("temporary_file/프리디그룹엘티디_배송자료 검수 리스트_10.05 (1).xlsx", sheet_name="Sheet1")
    sc = f.columns.get_loc('주문번호')
    print(sc)

    wb = openpyxl.load_workbook(filename="temporary_file/프리디그룹엘티디_배송자료 검수 리스트_10.05 (1).xlsx")
    ws = wb.worksheets[0]
    for row_idx, row in enumerate(f.values.tolist(), 2):
        print(row[sc])
        p = ProductPayment.objects.filter(secret=row[sc])
        if not p:
            ws.cell(row=row_idx, column=13).value = 'NO RESULT'
            ws.cell(row=row_idx, column=14).value = 'NO RESULT'
            ws.cell(row=row_idx, column=15).value = 'NO RESULT'
            continue
        p = p.first().product_bookings.first()

        if p.status == 5:
            ws.cell(row=row_idx, column=13).value = 'Transaction complete'
            ws.cell(row=row_idx, column=14).value = p.track_number
            ws.cell(row=row_idx, column=15).value = p.courier_company
        else:
            ws.cell(row=row_idx, column=13).value = 'Transaction not complete'
            ws.cell(row=row_idx, column=14).value = p.track_number
            ws.cell(row=row_idx, column=15).value = p.courier_company
        
    wb.save(filename="temporary_file/프리디그룹엘티디_배송자료 검수 리스트_10.05 (1).xlsx")
    wb.close()
